import os
from openpyxl import Workbook
from openpyxl import load_workbook

'''
This file defines some utility functions of opening and saving an excel file.
'''


# return the content of a single-sheet excel file in the form of a 2D list
def open_excel_single_sheet(directory, file_name):
    path = os.path.join(directory, '/{}.xlsx'.format(file_name))
    wb = load_workbook(path)
    booksheet = wb.active
    data = []
    for row in booksheet.rows:
        line = [col.value for col in row]
        data.append(line)
    return data


# # return the content of a single-sheet excel file in the form of a 2D list
# # the directory is the current directory
# def open_excel_single_sheet(file_name):
#     return open_excel_single_sheet('./', file_name)


# return the content of one sheet from a multi-sheet excel file by the sheet number
# or the sheet name
def open_excel_multi_sheets_one(directory, file_name, sheet_num=-1, sheet_name=''):
    if (sheet_num == -1 and sheet_name == '') or (sheet_name != -1 and sheet_name != ''):
        raise Exception("Invalid Argument")
    path = os.path.join(directory, '{}.xlsx'.format(file_name))
    wb = load_workbook(path)
    if sheet_num != -1:
        sheet_names = wb.sheetnames
        sheet_name = sheet_names[sheet_num]
        booksheet = wb[sheet_name]
    else:
        booksheet = wb[sheet_name]
    data = []
    for row in booksheet.rows:
        line = [col.value for col in row]
        data.append(line)
    return data


# # return the content of one sheet from a multi-sheet excel file by the sheet number
# # or the sheet name
# # the directory is the current directory
# def open_excel_multi_sheets_one(file_name, sheet_num=-1, sheet_name=''):
#     return open_excel_multi_sheets_one('.', file_name, sheet_num, sheet_name)


# return the content of one all sheets from a multi-sheet excel in the form of a dictionary
# indexed by its sheet name
def open_xlsx_file_multi_sheets_all(directory, file_name):
    path = os.path.join(directory, '{}.xlsx'.format(file_name))
    wb = load_workbook(path)
    sheet_names = wb.sheetnames
    data = {}
    for sheet_name in sheet_names:
        booksheet = wb[str(sheet_name)]
        _data = []
        for row in booksheet.rows:
            line = [col.value for col in row]
            _data.append(line)
        data[sheet_name] = _data
    return data


# def open_xlsx_file_multi_sheets_all(file_name):
#     return open_xlsx_file_multi_sheets_all('./', file_name)


# save a single-sheet excel file
def save_xlsx_file_single_sheet(directory, file_name, data):
    path = os.path.join(directory, '{}.xlsx'.format(file_name))
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    for row in data:
        ws.append(row)
    wb.save(path)


# # save a single-sheet excel file
# # the directory is the current directory
# def save_xlsx_file_single_sheet(file_name, data):
#     save_xlsx_file_single_sheet('./', file_name, data)


# save a multi-sheet excel file
def save_xlsx_file_multi_sheets(directory, file_name, data, sheet_names):
    path = os.path.join(directory, '{}.xlsx'.format(file_name))
    if len(data) != len(sheet_names):
        assert False
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)
    for i in range(len(data)):
        _data = data[i]
        sheet_name = sheet_names[i]
        wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        for row in _data:
            ws.append(row)
    wb.save(path)


# # save a single-sheet excel file
# # the directory is the current directory
# def save_xlsx_file_multi_sheets(file_name, data, sheet_names):
#     save_xlsx_file_multi_sheets('./', file_name, data, sheet_names)
#
